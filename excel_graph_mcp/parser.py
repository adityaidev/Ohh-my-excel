from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from excel_graph_mcp.formula_parser import parse_formula


class ParsedCell:
    def __init__(self, address: str, sheet: str, value: Any, formula: Optional[str] = None):
        self.address = address
        self.sheet = sheet
        self.value = value
        self.formula = formula
        self.analysis = parse_formula(formula) if formula else None

    @property
    def is_formula(self) -> bool:
        return self.formula is not None

    @property
    def is_input(self) -> bool:
        return not self.is_formula and self.value is not None

    def to_dict(self) -> dict:
        return {
            "address": self.address,
            "sheet": self.sheet,
            "value_type": type(self.value).__name__ if self.value is not None else "none",
            "has_formula": self.is_formula,
            "formula_text": self.formula,
            "is_input": self.is_input,
            "functions_used": self.analysis.functions_used if self.analysis else [],
            "nesting_depth": self.analysis.nesting_depth if self.analysis else 0,
            "references": self.analysis.all_references if self.analysis else [],
        }


class ParsedTable:
    def __init__(self, name: str, sheet: str, range_str: str, columns: list[str]):
        self.name = name
        self.sheet = sheet
        self.range = range_str
        self.columns = columns

    def to_dict(self) -> dict:
        return {"name": self.name, "sheet": self.sheet, "range": self.range, "columns": self.columns}


class ParsedChart:
    def __init__(self, name: str, chart_type: str, sheet: str, data_source: Optional[str] = None):
        self.name = name
        self.type = chart_type
        self.sheet = sheet
        self.data_source = data_source

    def to_dict(self) -> dict:
        return {"name": self.name, "type": self.type, "sheet": self.sheet, "data_source": self.data_source}


class ParsedSheet:
    def __init__(self, name: str, index: int):
        self.name = name
        self.index = index
        self.cells: list[ParsedCell] = []
        self.tables: list[ParsedTable] = []
        self.charts: list[ParsedChart] = []

    @property
    def formula_count(self) -> int:
        return sum(1 for c in self.cells if c.is_formula)

    @property
    def cell_count(self) -> int:
        return len(self.cells)

    def to_dict(self) -> dict:
        return {
            "name": self.name,
            "index": self.index,
            "cell_count": self.cell_count,
            "formula_count": self.formula_count,
            "tables": [t.to_dict() for t in self.tables],
            "charts": [c.to_dict() for c in self.charts],
        }


class ParsedWorkbook:
    def __init__(self, path: Path):
        self.path = path
        self.name = path.stem
        self.sheets: list[ParsedSheet] = []
        self.named_ranges: list[dict] = []
        self._parse()

    def _parse(self):
        wb: Workbook = load_workbook(self.path, data_only=False, read_only=True)
        for idx, ws_name in enumerate(wb.sheetnames):
            ws: Worksheet = wb[ws_name]
            sheet = ParsedSheet(ws_name, idx)

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        formula = None
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            formula = cell.value
                            value = None
                        else:
                            value = cell.value
                        sheet.cells.append(ParsedCell(cell.coordinate, ws_name, value, formula))

            if hasattr(ws, "tables"):
                for table in ws.tables:
                    columns = [col.name for col in (table.tableColumns if hasattr(table, "tableColumns") else [])]
                    sheet.tables.append(ParsedTable(table.name, ws_name, table.ref, columns))

            self.sheets.append(sheet)

        wb.close()

        for name, definition in (wb.defined_names or {}).items():
            self.named_ranges.append({"name": name, "range": str(definition)})

    @property
    def sheet_count(self) -> int:
        return len(self.sheets)

    @property
    def size_bytes(self) -> int:
        return self.path.stat().st_size

    def to_dict(self) -> dict:
        return {
            "path": str(self.path),
            "name": self.name,
            "sheet_count": self.sheet_count,
            "size_bytes": self.size_bytes,
            "sheets": [s.to_dict() for s in self.sheets],
            "named_ranges": self.named_ranges,
        }
