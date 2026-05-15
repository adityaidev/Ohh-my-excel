from pathlib import Path


def generate_workbook(prompt: str, output_path: str, template: str = "", data_source: str = "", style: str = "professional") -> dict:
    p = Path(output_path)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Generated from prompt"
    ws["A2"] = prompt[:100]
    wb.save(str(p))
    from excel_graph_mcp.dependency import build_dependency_graph
    store, builder = build_dependency_graph(p)
    stats = store.stats()
    store.close()
    return {"file": str(p), "status": "created", "sheets": 1, "cells": 2, "nodes": stats["nodes"], "edges": stats["edges"]}


def generate_workbook_from_data(data: str, output_path: str, sheet_name: str = "Sheet1", include_charts: bool = False, style: str = "professional") -> dict:
    p = Path(output_path)
    import csv
    import io
    import json

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    try:
        rows = json.loads(data) if isinstance(data, str) else data
        if isinstance(rows, list) and rows:
            headers = list(rows[0].keys())
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
            for row_idx, row in enumerate(rows, 2):
                for col_idx, h in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))
    except (json.JSONDecodeError, TypeError):
        reader = csv.DictReader(io.StringIO(data))
        for col, h in enumerate(reader.fieldnames or [], 1):
            ws.cell(row=1, column=col, value=h)
        for row_idx, row in enumerate(reader, 2):
            for col_idx, h in enumerate(reader.fieldnames or [], 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))
    wb.save(str(p))
    from excel_graph_mcp.dependency import build_dependency_graph
    store, builder = build_dependency_graph(p)
    stats = store.stats()
    store.close()
    return {"file": str(p), "status": "created", "sheets": 1, "nodes": stats["nodes"]}


TEMPLATES = {
    "expense_tracker": {"sheets": [{"name": "Expenses", "columns": ["Date", "Category", "Description", "Amount"]}, {"name": "Summary", "columns": ["Category", "Total"]}]},
    "budget_planner": {"sheets": [{"name": "Budget", "columns": ["Category", "Budgeted", "Actual", "Difference"]}]},
    "invoice": {"sheets": [{"name": "Invoice", "columns": ["Item", "Quantity", "Unit Price", "Total"]}]},
}


def generate_workbook_from_template(template_name: str, output_path: str, customizations: str = "") -> dict:
    p = Path(output_path)
    template = TEMPLATES.get(template_name)
    if not template:
        return {"error": f"Template '{template_name}' not found. Available: {list(TEMPLATES.keys())}"}
    from openpyxl import Workbook
    wb = Workbook()
    for sheet_spec in template["sheets"]:
        ws = wb.active if sheet_spec == template["sheets"][0] else wb.create_sheet()
        ws.title = sheet_spec["name"]
        for col, h in enumerate(sheet_spec["columns"], 1):
            ws.cell(row=1, column=col, value=h)
    wb.save(str(p))
    return {"file": str(p), "status": "created", "template": template_name}


def add_sheet(file_path: str, sheet_name: str, columns: list[dict], formulas: list[dict] = None, formatting: dict = None) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    ws = wb.create_sheet(title=sheet_name)
    for col, col_def in enumerate(columns, 1):
        ws.cell(row=1, column=col, value=col_def.get("name", f"Column{col}"))
    if formulas:
        for f in formulas:
            ws.cell(row=int(f.get("row", 2)), column=ord(f.get("cell", "A")[0]) - 64, value=f.get("formula", ""))
    wb.save(file_path)
    from excel_graph_mcp.dependency import build_dependency_graph
    store, builder = build_dependency_graph(Path(file_path))
    store.close()
    return {"file": file_path, "status": "sheet_added", "sheet": sheet_name}


def add_chart(file_path: str, sheet_name: str, chart_type: str, data_range: str, title: str = "") -> dict:
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, LineChart, PieChart
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    chart_map = {"bar": BarChart, "line": LineChart, "pie": PieChart}
    chart_cls = chart_map.get(chart_type, BarChart)
    chart = chart_cls()
    chart.title = title or f"{chart_type} chart"
    ws.add_chart(chart, "E5")
    wb.save(file_path)
    return {"file": file_path, "status": "chart_added", "type": chart_type}


def apply_formatting(file_path: str, style: str = "professional", auto_size: bool = True, freeze_panes: dict = None) -> dict:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    wb = load_workbook(file_path)
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        if auto_size:
            for col in ws.columns:
                max_length = max((len(str(cell.value or "")) for cell in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    wb.save(file_path)
    return {"file": file_path, "status": "formatted", "style": style}


def generate_formulas(file_path: str, sheet_name: str, column: str, formula_type: str = "sum") -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    max_row = ws.max_row or 1
    formula_map = {
        "sum": f"=SUM({column}2:{column}{max_row})",
        "average": f"=AVERAGE({column}2:{column}{max_row})",
        "count": f"=COUNTA({column}2:{column}{max_row})",
        "max": f"=MAX({column}2:{column}{max_row})",
        "min": f"=MIN({column}2:{column}{max_row})",
    }
    formula = formula_map.get(formula_type, formula_map["sum"])
    target_cell = f"{column}{max_row + 2}"
    ws[target_cell] = formula
    wb.save(file_path)
    return {"file": file_path, "sheet": sheet_name, "target": target_cell, "formula": formula, "type": formula_type}


def validate_workbook(file_path: str, check_circular: bool = True, check_broken_refs: bool = True) -> dict:
    from excel_graph_mcp.dependency import build_dependency_graph
    store, builder = build_dependency_graph(Path(file_path))
    issues = []
    if check_circular:
        cycles = builder.find_circular_references()
        if cycles:
            issues.append({"type": "circular_reference", "count": len(cycles), "details": cycles[:5]})
    if check_broken_refs:
        broken = builder.find_broken_references()
        if broken:
            issues.append({"type": "broken_reference", "count": len(broken), "details": broken[:10]})
    store.close()
    return {"file": file_path, "valid": len(issues) == 0, "issues": issues}
