from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

FIXTURE_DIR = Path(__file__).parent


def generate_simple_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Item"
    ws["B1"] = "Price"
    ws["C1"] = "Quantity"
    ws["D1"] = "Total"
    ws["A2"] = "Widget"
    ws["B2"] = 10
    ws["C2"] = 5
    ws["D2"] = "=B2*C2"
    ws["A3"] = "Gadget"
    ws["B3"] = 25
    ws["C3"] = 3
    ws["D3"] = "=B3*C3"
    ws["D4"] = "=SUM(D2:D3)"
    wb.save(str(FIXTURE_DIR / "simple.xlsx"))


def generate_cross_sheet_workbook():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Input"
    ws1["A1"] = 100
    ws1["A2"] = 200
    ws1["A3"] = 300
    ws2 = wb.create_sheet("Calculation")
    ws2["A1"] = "=Input!A1*2"
    ws2["A2"] = "=Input!A2+Input!A3"
    ws2["A3"] = "=SUM(Input!A1:A3)"
    wb.save(str(FIXTURE_DIR / "cross_sheet.xlsx"))


def generate_workbook_with_tables():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    headers = ["Name", "Department", "Salary"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    data = [
        ["Alice", "Engineering", 120000],
        ["Bob", "Marketing", 90000],
        ["Charlie", "Engineering", 110000],
    ]
    for row_idx, row in enumerate(data, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    ws.cell(row=5, column=3, value="=SUM(C2:C4)")
    tab = Table(displayName="EmployeeTable", ref="A1:C4")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)
    wb.save(str(FIXTURE_DIR / "with_tables.xlsx"))


def generate_all():
    generate_simple_workbook()
    generate_cross_sheet_workbook()
    generate_workbook_with_tables()
    print("Fixtures generated in", FIXTURE_DIR)


if __name__ == "__main__":
    generate_all()
